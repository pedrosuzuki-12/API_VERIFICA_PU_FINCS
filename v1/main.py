import logging
import pandas as pd
from openpyxl import load_workbook
import os

import config
from utils.formatters import num_br
from scrapers.vortx import (
    busca_id_vortx,
    busca_historico_vortx,
    checagem_vortx,
)
from scrapers.oliveira_trust import (
    busca_id_oliveira,
    busca_historico_oliveira,
    checagem_oliveira,
)

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler('automation.log')
    ]
)
logger = logging.getLogger(__name__)

def process_file():
    arquivo_input = config.ARQUIVO_INPUT
    aba = config.ABA_EXCEL
    col_retorno = config.COLUNA_RETORNO

    if not os.path.exists(arquivo_input):
        logger.error(f"Input file not found: {arquivo_input}")
        return

    logger.info(f"Starting value review automation using {arquivo_input}, sheet {aba}")

    df = pd.read_excel(arquivo_input, sheet_name=aba, usecols="A:H")
    df = df.dropna(subset=[df.columns[0]])

    wb = load_workbook(arquivo_input)
    ws = wb[aba]

    for index, row in df.iterrows():
        linha_excel = index + 2

        codigo_if = str(row.iloc[0]).strip()
        data_evento = pd.to_datetime(row.iloc[1]).strftime('%Y-%m-%d')
        amort_ord_per = num_br(row.iloc[2])
        amex_per = num_br(row.iloc[3])
        juros_per = num_br(row.iloc[4])
        incorp_per = num_br(row.iloc[6])
        af = str(row.iloc[7]).strip().upper()
        
        logger.info(f"Processing row {linha_excel}: Código IF {codigo_if}, AF {af}")

        if af == "OLIVEIRA TRUST DTVM S.A.":
            res_id = busca_id_oliveira(codigo_if)
            if not res_id:
                logger.warning(f"Título {codigo_if} não encontrado na Oliveira Trust.")
                ws[f"{col_retorno}{linha_excel}"] = "Título não encontrado na Oliveira Trust"
                continue

            id_ot, nome_titulo = res_id
            logger.info(f"Título encontrado: {nome_titulo} (código {codigo_if})")

            dados_evento = busca_historico_oliveira(id_ot, data_evento)

            if dados_evento is None:
                logger.warning(f"Nenhum evento encontrado na data {data_evento}.")
                ws[f"{col_retorno}{linha_excel}"] = "Nenhum evento encontrado na data"
            else:
                vn, juros, pu, amort_pgto, juros_pgto, premio_pgto, total_pgto = dados_evento
                
                resultado = checagem_oliveira(
                    amort_ord_per=amort_ord_per,
                    amex_per=amex_per,
                    incorp_per=incorp_per,
                    vn=vn,
                    juros=juros,
                    pu=pu,
                    amort_pgto=amort_pgto,
                    juros_pgto=juros_pgto,
                    premio_pgto=premio_pgto,
                    total_pgto=total_pgto,
                    juros_per=juros_per
                )
        
                logger.info(f"Log Resumo: {resultado['log_suscinto']}")
                ws[f"{col_retorno}{linha_excel}"] = resultado["log_suscinto"]

        elif af == "VORTX DTVM LTDA.":
            res_id = busca_id_vortx(codigo_if)
            if not res_id:
                logger.warning(f"Título {codigo_if} não encontrado na Vórtx.")
                ws[f"{col_retorno}{linha_excel}"] = "Título não encontrado na Vórtx"
                continue 

            id_vortx = res_id
            logger.info(f"Título encontrado: (código {codigo_if})")

            dados_evento = busca_historico_vortx(id_vortx, data_evento)

            if dados_evento is None:
                logger.warning(f"Nenhum evento encontrado no histórico para a data {data_evento}.")
                ws[f"{col_retorno}{linha_excel}"] = "Nenhum evento encontrado na data"
                continue

            resultado = checagem_vortx(
                amort_ord_per=amort_ord_per,
                amex_per=amex_per,
                juros_per=juros_per,
                incorp_per=incorp_per,
                linha=dados_evento
            )

            logger.info(f"Log Resumo: {resultado['log_suscinto']}")
            ws[f"{col_retorno}{linha_excel}"] = resultado["log_suscinto"]

        else:
            logger.warning(f"AF desconhecido: {af}. Pulando.")
            ws[f"{col_retorno}{linha_excel}"] = f"AF não mapeado: {af}"

    nome_base = arquivo_input.rsplit('.', 1)[0]
    arquivo_saida = f"{nome_base}_Retorno.xlsx"

    wb.save(arquivo_saida)
    logger.info(f"Coluna Retorno atualizada e salva em: {arquivo_saida}")

if __name__ == "__main__":
    process_file()